import streamlit as st
import os
import shutil
import re
from pypdf import PdfReader
from docx import Document
from PIL import Image
import pytesseract
from openpyxl import Workbook, load_workbook
from io import BytesIO

# Set Tesseract path for Windows
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# ----------- Text Extraction Helpers -----------
def extract_text_from_pdf(file):
    try:
        reader = PdfReader(file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
        return text.lower()
    except Exception:
        return ""

def extract_text_from_docx(file):
    try:
        doc = Document(file)
        full_text = [para.text for para in doc.paragraphs]
        return "\n".join(full_text).lower()
    except Exception:
        return ""

def extract_text_from_image(file):
    try:
        img = Image.open(file)
        text = pytesseract.image_to_string(img)
        return text.lower()
    except Exception:
        return ""

def extract_text(file, ext):
    if ext == ".pdf":
        return extract_text_from_pdf(file)
    elif ext == ".docx":
        return extract_text_from_docx(file)
    elif ext in [".jpeg", ".jpg", ".png"]:
        return extract_text_from_image(file)
    else:
        return ""

# ----------- Matching Helpers -----------
def match_keywords(text, keywords):
    for kw in keywords:
        if kw.lower() in text:
            return True
    return False

def match_experience(text, experience_input):
    match = re.search(r'(\d+)\s*(\w+)', experience_input.lower())
    if not match:
        return False
    min_years = int(match.group(1))
    unit = match.group(2)
    pattern = re.compile(r'(\d+)\s*' + re.escape(unit))
    for m in pattern.finditer(text):
        if int(m.group(1)) >= min_years:
            return True
    return False

# ----------- Email & Phone Extraction -----------
def extract_emails(text):
    pattern = r'[\w\.-]+@(?:gmail\.com|yahoo\.com)'
    emails = re.findall(pattern, text, flags=re.IGNORECASE)
    return list(set(emails))

def extract_phone_numbers(text):
    cleaned_text = re.sub(r'[\s\-]', '', text)
    phones = []
    plus62_pattern = re.compile(r'\+62(\d{8,14})')
    phones += ['+62' + m for m in plus62_pattern.findall(cleaned_text)]
    literal_62plus_pattern = re.compile(r'62\+(\d{8,14})')
    phones += ['62+' + m for m in literal_62plus_pattern.findall(cleaned_text)]
    zero8_pattern = re.compile(r'08(\d{8,14})')
    phones += ['08' + m for m in zero8_pattern.findall(cleaned_text)]
    return list(set(phones))

# ----------- Streamlit App -----------
st.title("Resume Vetting Tool")

# Password Check
password = st.text_input("Masukkan kata sandi", type="password")
if password != "your_team_password":  # Replace with your desired password
    st.error("Kata sandi salah!")
    st.stop()

# Input fields
default_keywords = {
    "lokasi": "semarang",
    "pendidikan": "univ,smk,sma",
    "jurusan": "kearsipan,perpustakaan,administrasi",
    "perusahaan": "indoraj,arsip",
    "pengalaman": "1 tahun,satu tahun"
}

st.write("Masukkan kriteria vetting:")
lokasi = st.text_input("Lokasi (pisahkan koma)", default_keywords["lokasi"])
pendidikan = st.text_input("Pendidikan (pisahkan koma)", default_keywords["pendidikan"])
jurusan = st.text_input("Jurusan (pisahkan koma)", default_keywords["jurusan"])
perusahaan = st.text_input("Pengalaman Perusahaan (pisahkan koma)", default_keywords["perusahaan"])
pengalaman = st.text_input("Minimal Tahun Pengalaman (contoh: 5 tahun)", default_keywords["pengalaman"])
uploaded_files = st.file_uploader(
    "Pilih file CV (pdf, docx, jpg, png)",
    accept_multiple_files=True,
    type=["pdf", "docx", "jpg", "png"]
)

if st.button("Jalankan Vetting"):
    if not uploaded_files:
        st.error("⚠️ Tidak ada file dipilih!")
    else:
        # Define output folder (server-side, temporary)
        output_folder = "output"
        os.makedirs(os.path.join(output_folder, "Bagus_Sekali"), exist_ok=True)
        os.makedirs(os.path.join(output_folder, "Layak"), exist_ok=True)
        excel_path = os.path.join(output_folder, "Hasil_Vetting.xlsx")

        # Process inputs
        lokasi_keywords = [k.strip().lower() for k in lokasi.split(",") if k.strip()]
        pendidikan_keywords = [k.strip().lower() for k in pendidikan.split(",") if k.strip()]
        jurusan_keywords = [k.strip().lower() for k in jurusan.split(",") if k.strip()]
        perusahaan_keywords = [k.strip().lower() for k in perusahaan.split(",") if k.strip()]
        pengalaman_input = pengalaman.strip()

        weights = {
            "lokasi": 10,
            "pendidikan": 20,
            "jurusan": 20,
            "pengalaman_perusahaan": 20,
            "tahun_pengalaman": 10,
        }

        # Initialize Excel workbook
        wb = Workbook()
        ws_bagus = wb.create_sheet("Bagus_Sekali")
        ws_layak = wb.create_sheet("Layak")
        ws_bagus.append(["Nama File", "Email", "Nomor Telepon"])
        ws_layak.append(["Nama File", "Email", "Nomor Telepon"])

        # Process each uploaded file
        st.write("Hasil Vetting:")
        for file in uploaded_files:
            try:
                ext = os.path.splitext(file.name)[1].lower()
                text = extract_text(file, ext)
                score = 0

                if match_keywords(text, lokasi_keywords):
                    score += weights["lokasi"]
                if match_keywords(text, pendidikan_keywords):
                    score += weights["pendidikan"]
                if match_keywords(text, jurusan_keywords):
                    score += weights["jurusan"]
                if match_keywords(text, perusahaan_keywords):
                    score += weights["pengalaman_perusahaan"]
                if pengalaman_input and match_experience(text, pengalaman_input):
                    score += weights["tahun_pengalaman"]

                emails = extract_emails(text)
                phones = extract_phone_numbers(text)
                email_str = ", ".join(emails) if emails else "Tidak ditemukan"
                phone_str = ", ".join(phones) if phones else "Tidak ditemukan"

                if score >= 60:
                    target_folder = os.path.join(output_folder, "Bagus_Sekali")
                    ws = ws_bagus
                    category = "BAGUS SEKALI"
                elif score >= 30:
                    target_folder = os.path.join(output_folder, "Layak")
                    ws = ws_layak
                    category = "LAYAK"
                else:
                    st.write(f"❌ {file.name} - SKOR {score} (Ditolak)")
                    continue

                # Save file to target folder
                with open(os.path.join(target_folder, file.name), "wb") as f:
                    f.write(file.read())
                ws.append([file.name, email_str, phone_str])
                st.write(f"✅ {file.name} - SKOR {score} ({category})")

            except Exception as e:
                st.error(f"⚠️ Error processing {file.name}: {e}")

        # Save Excel file and offer download
        wb.save(excel_path)
        with open(excel_path, "rb") as f:
            st.download_button(
                label="Unduh Hasil Vetting",
                data=f,
                file_name="Hasil_Vetting.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.success(f"💾 Berhasil menyimpan hasil ke {excel_path}")